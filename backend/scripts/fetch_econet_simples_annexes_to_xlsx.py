from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from backend.app.core.config import get_settings  # noqa: E402
from backend.app.services.integrations.econet.assisted_session import get_econet_assisted_session  # noqa: E402
from backend.app.services.integrations.econet.client import EconetClient  # noqa: E402
from backend.app.services.integrations.econet.errors import EconetSessionError  # noqa: E402
from backend.app.services.integrations.econet.parser import (  # noqa: E402
    format_cnae,
    normalize_cnae,
    parse_search_results,
    parse_simples_nacional,
)
from backend.scripts.export_econet_simples_annex_audit import ANNEX_HEADERS, CATALOG_SHEET_NAME  # noqa: E402


FINAL_OK_STATUSES = {"OK", "PROHIBITED"}


@dataclass(slots=True)
class TargetRow:
    row_idx: int
    cnae: str


@dataclass(slots=True)
class QueryResult:
    annex_default: str | None
    annex_conditional: str | None
    simples_status: str | None
    simples_allowed: bool | None
    annex_status: str
    note: str | None


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Consulta a aba de Simples Nacional da Econet e preenche anexos no XLSX.")
    parser.add_argument("--input-xlsx", type=Path, required=True, help="Planilha XLSX a atualizar.")
    parser.add_argument("--output-xlsx", type=Path, required=False, help="Arquivo XLSX de saida.")
    parser.add_argument("--sheet-name", type=str, default=CATALOG_SHEET_NAME, help="Aba do catalogo.")
    parser.add_argument("--session-file", type=Path, default=Path("backend/storage/sessions/econet/manual-storage-state.json"))
    parser.add_argument("--batch-size", type=int, default=50)
    parser.add_argument("--query-mode", choices=("missing_only", "all"), default="missing_only")
    parser.add_argument("--skip-probe", action="store_true")
    return parser


def _ensure_annex_headers(sheet) -> dict[str, int]:
    header_cells = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    header_map = {str(value).strip(): index + 1 for index, value in enumerate(header_cells) if value is not None}
    cnae_column = header_map.get("CNAE Normalizado") or header_map.get("CNAE")
    if cnae_column is None:
        raise ValueError("Worksheet must contain header 'CNAE Normalizado' or 'CNAE'.")

    next_col = sheet.max_column + 1
    for header in ANNEX_HEADERS:
        if header not in header_map:
            sheet.cell(row=1, column=next_col).value = header
            header_map[header] = next_col
            next_col += 1
    header_map["__cnae__"] = cnae_column
    return header_map


def _select_target_rows(sheet, *, header_map: dict[str, int], query_mode: str) -> list[TargetRow]:
    cnae_col = header_map["__cnae__"]
    status_col = header_map["anexo_econet_status"]
    rows: list[TargetRow] = []
    for row_idx in range(2, sheet.max_row + 1):
        raw_cnae = sheet.cell(row=row_idx, column=cnae_col).value
        if raw_cnae is None:
            continue
        cnae = normalize_cnae(str(raw_cnae))
        if query_mode == "missing_only":
            current_status = sheet.cell(row=row_idx, column=status_col).value
            if str(current_status).strip().upper() in FINAL_OK_STATUSES:
                continue
        rows.append(TargetRow(row_idx=row_idx, cnae=cnae))
    return rows


def _load_storage_state(session_file: Path) -> dict:
    payload = json.loads(session_file.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("Session file must contain a JSON object.")
    return payload


def _import_session(session_file: Path) -> EconetClient:
    settings = get_settings().model_copy(update={"econet_assisted_session_enabled": True})
    assisted_session = get_econet_assisted_session(settings)
    assisted_session.import_storage_state(_load_storage_state(session_file))
    client = EconetClient(settings=settings)
    return client


def _search_exact_econet_id(client: EconetClient, cnae: str) -> str:
    results = parse_search_results(client.search_cnae(cnae))
    exact = [item for item in results if item.cnae == cnae]
    if not exact:
        raise ValueError(f"CNAE {format_cnae(cnae)} not found in Econet search.")
    if len(exact) > 1:
        raise ValueError(f"CNAE {format_cnae(cnae)} returned multiple exact matches in Econet search.")
    return exact[0].econet_id_cnae


def _query_simples_for_cnae(client: EconetClient, cnae: str) -> QueryResult:
    try:
        econet_id = _search_exact_econet_id(client, cnae)
        simples = parse_simples_nacional(client.get_simples_nacional(econet_id))
        annex_status = "OK"
        if simples.status == "PROHIBITED":
            annex_status = "PROHIBITED"
        elif simples.annex_default is None:
            annex_status = "NO_ANNEX"
        return QueryResult(
            annex_default=simples.annex_default,
            annex_conditional=simples.annex_conditional,
            simples_status=simples.status,
            simples_allowed=simples.allowed,
            annex_status=annex_status,
            note=None,
        )
    except EconetSessionError:
        raise
    except Exception as exc:
        return QueryResult(
            annex_default=None,
            annex_conditional=None,
            simples_status=None,
            simples_allowed=None,
            annex_status="QUERY_ERROR",
            note=str(exc),
        )


def _write_result(sheet, *, header_map: dict[str, int], target_row: TargetRow, result: QueryResult) -> None:
    values = {
        "anexo_econet_padrao": result.annex_default,
        "anexo_econet_condicional": result.annex_conditional,
        "simples_status_econet": result.simples_status,
        "simples_permitido_econet": result.simples_allowed,
        "anexo_econet_status": result.annex_status,
        "anexo_econet_observacao": result.note,
        "cache_retrieved_at_econet": None,
        "cache_expires_at_econet": None,
    }
    for header, value in values.items():
        sheet.cell(row=target_row.row_idx, column=header_map[header]).value = value


def run_fetch(
    *,
    input_xlsx: Path,
    output_xlsx: Path,
    session_file: Path,
    batch_size: int,
    sheet_name: str,
    query_mode: str,
    skip_probe: bool,
) -> dict[str, int]:
    workbook = load_workbook(input_xlsx)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet '{sheet_name}' was not found in '{input_xlsx}'.")
    sheet = workbook[sheet_name]
    header_map = _ensure_annex_headers(sheet)
    targets = _select_target_rows(sheet, header_map=header_map, query_mode=query_mode)

    queried = 0
    ok = 0
    prohibited = 0
    no_annex = 0
    query_error = 0

    client = _import_session(session_file)
    try:
        if not skip_probe:
            client.probe_session()

        for offset in range(0, len(targets), batch_size):
            batch = targets[offset : offset + batch_size]
            print(
                json.dumps(
                    {
                        "event": "batch_start",
                        "batch_number": (offset // batch_size) + 1,
                        "batch_size": len(batch),
                        "start": offset + 1,
                        "end": offset + len(batch),
                        "total": len(targets),
                    },
                    ensure_ascii=True,
                )
            )
            for target in batch:
                result = _query_simples_for_cnae(client, target.cnae)
                _write_result(sheet, header_map=header_map, target_row=target, result=result)
                queried += 1
                if result.annex_status == "OK":
                    ok += 1
                elif result.annex_status == "PROHIBITED":
                    prohibited += 1
                elif result.annex_status == "NO_ANNEX":
                    no_annex += 1
                elif result.annex_status == "QUERY_ERROR":
                    query_error += 1
            workbook.save(output_xlsx)
            print(
                json.dumps(
                    {
                        "event": "batch_saved",
                        "batch_number": (offset // batch_size) + 1,
                        "queried": queried,
                        "ok": ok,
                        "prohibited": prohibited,
                        "no_annex": no_annex,
                        "query_error": query_error,
                        "output_xlsx": str(output_xlsx),
                    },
                    ensure_ascii=True,
                )
            )
    finally:
        client.close()

    return {
        "target_rows": len(targets),
        "queried": queried,
        "ok": ok,
        "prohibited": prohibited,
        "no_annex": no_annex,
        "query_error": query_error,
    }


def _resolve_output_path(input_xlsx: Path, output_xlsx: Path | None) -> Path:
    if output_xlsx is not None:
        return output_xlsx
    return input_xlsx.with_name(f"{input_xlsx.stem}_consultado_econet.xlsx")


def main() -> None:
    args = build_parser().parse_args()
    output_xlsx = _resolve_output_path(args.input_xlsx, args.output_xlsx)
    summary = run_fetch(
        input_xlsx=args.input_xlsx,
        output_xlsx=output_xlsx,
        session_file=args.session_file,
        batch_size=args.batch_size,
        sheet_name=args.sheet_name,
        query_mode=args.query_mode,
        skip_probe=args.skip_probe,
    )
    print(json.dumps({"output_xlsx": str(output_xlsx), "summary": summary}, ensure_ascii=True))


if __name__ == "__main__":
    main()
