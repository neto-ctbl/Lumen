from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from backend.app.db.session import SessionLocal  # noqa: E402
from backend.app.models.econet_cnae_cache import EconetCnaeCache  # noqa: E402
from backend.app.services.integrations.econet.parser import format_cnae, normalize_cnae  # noqa: E402


CATALOG_SHEET_NAME = "Catalogo"
SUMMARY_SHEET_NAME = "ResumoAnexos"
ANNEX_HEADERS = (
    "anexo_econet_padrao",
    "anexo_econet_condicional",
    "simples_status_econet",
    "simples_permitido_econet",
    "anexo_econet_status",
    "anexo_econet_observacao",
    "cache_retrieved_at_econet",
    "cache_expires_at_econet",
)


@dataclass(slots=True)
class AnnexAuditRow:
    cnae: str
    cnae_normalized: str
    annex_default: str | None
    annex_conditional: str | None
    simples_status: str | None
    simples_allowed: bool | None
    annex_status: str
    cache_retrieved_at: datetime | None
    cache_expires_at: datetime | None


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Exporta anexos do Simples da Econet para a planilha de catalogo CNAE.")
    parser.add_argument("--input-xlsx", type=Path, required=False, help="Planilha XLSX do catalogo CNAE.")
    parser.add_argument("--input-json", type=Path, required=False, help="Catalogo JSON materializado.")
    parser.add_argument("--output-xlsx", type=Path, required=False, help="Arquivo XLSX de saida.")
    parser.add_argument("--sheet-name", type=str, default=CATALOG_SHEET_NAME, help="Aba do XLSX a enriquecer.")
    return parser


def load_cache_map(session: Session, *, cnaes: list[str]) -> dict[str, EconetCnaeCache]:
    if not cnaes:
        return {}
    rows = session.scalars(select(EconetCnaeCache).where(EconetCnaeCache.cnae.in_(cnaes))).all()
    return {row.cnae: row for row in rows}


def build_annex_audit(cache: EconetCnaeCache | None, *, cnae_normalized: str) -> AnnexAuditRow:
    if cache is None:
        return AnnexAuditRow(
            cnae=format_cnae(cnae_normalized),
            cnae_normalized=cnae_normalized,
            annex_default=None,
            annex_conditional=None,
            simples_status=None,
            simples_allowed=None,
            annex_status="MISSING_CACHE",
            cache_retrieved_at=None,
            cache_expires_at=None,
        )

    annex_status = "OK"
    if cache.simples_status == "PROHIBITED":
        annex_status = "PROHIBITED"
    elif cache.simples_annex_default is None:
        annex_status = "NO_ANNEX"

    return AnnexAuditRow(
        cnae=cache.cnae_formatted,
        cnae_normalized=cache.cnae,
        annex_default=cache.simples_annex_default,
        annex_conditional=cache.simples_annex_conditional,
        simples_status=cache.simples_status,
        simples_allowed=cache.simples_allowed,
        annex_status=annex_status,
        cache_retrieved_at=cache.retrieved_at,
        cache_expires_at=cache.expires_at,
    )


def summarize_audit(rows: list[AnnexAuditRow]) -> dict[str, int]:
    summary = {
        "rows_total": len(rows),
        "ok": 0,
        "prohibited": 0,
        "no_annex": 0,
        "missing_cache": 0,
    }
    for row in rows:
        if row.annex_status == "OK":
            summary["ok"] += 1
        elif row.annex_status == "PROHIBITED":
            summary["prohibited"] += 1
        elif row.annex_status == "NO_ANNEX":
            summary["no_annex"] += 1
        elif row.annex_status == "MISSING_CACHE":
            summary["missing_cache"] += 1
    return summary


def _normalize_cnae_cell(value: Any) -> str:
    if value is None:
        raise ValueError("Linha sem CNAE.")
    return normalize_cnae(str(value))


def _append_summary_sheet(workbook: Workbook, *, rows: list[AnnexAuditRow]) -> None:
    if SUMMARY_SHEET_NAME in workbook.sheetnames:
        del workbook[SUMMARY_SHEET_NAME]
    sheet = workbook.create_sheet(SUMMARY_SHEET_NAME)
    summary = summarize_audit(rows)
    sheet.append(("metrica", "valor"))
    for key, value in summary.items():
        sheet.append((key, value))

    sheet.append(())
    sheet.append(("cnae", "cnae_normalizado", "anexo_econet_status"))
    for row in rows:
        if row.annex_status == "MISSING_CACHE":
            sheet.append((row.cnae, row.cnae_normalized, row.annex_status))


def export_from_xlsx(
    session: Session,
    *,
    input_path: Path,
    output_path: Path,
    sheet_name: str = CATALOG_SHEET_NAME,
) -> dict[str, int]:
    workbook = load_workbook(input_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet '{sheet_name}' was not found in '{input_path}'.")
    sheet = workbook[sheet_name]
    header_cells = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    header_map = {str(value).strip(): index + 1 for index, value in enumerate(header_cells) if value is not None}

    cnae_column = header_map.get("CNAE Normalizado") or header_map.get("CNAE")
    if cnae_column is None:
        raise ValueError("Worksheet must contain header 'CNAE Normalizado' or 'CNAE'.")

    cnaes: list[str] = []
    for row_idx in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row_idx, column=cnae_column).value
        if value is None:
            continue
        cnaes.append(_normalize_cnae_cell(value))

    cache_map = load_cache_map(session, cnaes=cnaes)
    audit_rows: list[AnnexAuditRow] = []
    start_column = sheet.max_column + 1
    for offset, header in enumerate(ANNEX_HEADERS):
        sheet.cell(row=1, column=start_column + offset).value = header

    for row_idx in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row_idx, column=cnae_column).value
        if value is None:
            continue
        cnae_normalized = _normalize_cnae_cell(value)
        audit = build_annex_audit(cache_map.get(cnae_normalized), cnae_normalized=cnae_normalized)
        audit_rows.append(audit)
        row_values = (
            audit.annex_default,
            audit.annex_conditional,
            audit.simples_status,
            audit.simples_allowed,
            audit.annex_status,
            None,
            audit.cache_retrieved_at.isoformat() if audit.cache_retrieved_at else None,
            audit.cache_expires_at.isoformat() if audit.cache_expires_at else None,
        )
        for offset, cell_value in enumerate(row_values):
            sheet.cell(row=row_idx, column=start_column + offset).value = cell_value

    _append_summary_sheet(workbook, rows=audit_rows)
    workbook.save(output_path)
    return summarize_audit(audit_rows)


def export_from_json(
    session: Session,
    *,
    input_path: Path,
    output_path: Path,
) -> dict[str, int]:
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    rows = payload.get("catalog")
    if not isinstance(rows, list):
        raise ValueError("Input JSON must contain a top-level 'catalog' list.")

    cnaes = [_normalize_cnae_cell(item.get("normalized") or item.get("cnae")) for item in rows]
    cache_map = load_cache_map(session, cnaes=cnaes)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = CATALOG_SHEET_NAME
    headers = (
        "CNAE",
        "CNAE Normalizado",
        "Divisao",
        "Denominacao",
        "activity_type",
        "rule_id",
        *ANNEX_HEADERS,
    )
    sheet.append(headers)

    audit_rows: list[AnnexAuditRow] = []
    for item, cnae_normalized in zip(rows, cnaes, strict=False):
        audit = build_annex_audit(cache_map.get(cnae_normalized), cnae_normalized=cnae_normalized)
        audit_rows.append(audit)
        sheet.append(
            (
                item.get("cnae"),
                cnae_normalized,
                item.get("division"),
                item.get("denomination"),
                item.get("activity_type"),
                item.get("rule_id"),
                audit.annex_default,
                audit.annex_conditional,
                audit.simples_status,
                audit.simples_allowed,
                audit.annex_status,
                None,
                audit.cache_retrieved_at.isoformat() if audit.cache_retrieved_at else None,
                audit.cache_expires_at.isoformat() if audit.cache_expires_at else None,
            )
        )

    _append_summary_sheet(workbook, rows=audit_rows)
    workbook.save(output_path)
    return summarize_audit(audit_rows)


def _resolve_output_path(*, input_path: Path | None, explicit_output: Path | None) -> Path:
    if explicit_output is not None:
        return explicit_output
    if input_path is None:
        raise ValueError("output_xlsx is required when there is no input file path to derive from.")
    return input_path.with_name(f"{input_path.stem}_com_anexos_econet.xlsx")


def main() -> None:
    args = build_parser().parse_args()
    if bool(args.input_xlsx) == bool(args.input_json):
        raise SystemExit("Provide exactly one of --input-xlsx or --input-json.")

    session = SessionLocal()
    try:
        source_path = args.input_xlsx or args.input_json
        output_path = _resolve_output_path(input_path=source_path, explicit_output=args.output_xlsx)
        if args.input_xlsx is not None:
            summary = export_from_xlsx(
                session,
                input_path=args.input_xlsx,
                output_path=output_path,
                sheet_name=args.sheet_name,
            )
        else:
            summary = export_from_json(
                session,
                input_path=args.input_json,
                output_path=output_path,
            )
        print(json.dumps({"output_xlsx": str(output_path), "summary": summary}, ensure_ascii=True))
    finally:
        session.close()


if __name__ == "__main__":
    main()
