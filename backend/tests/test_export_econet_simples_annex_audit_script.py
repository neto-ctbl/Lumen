from __future__ import annotations

import json
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

from backend.app.models.econet_cnae_cache import EconetCnaeCache
from backend.scripts.export_econet_simples_annex_audit import (
    build_annex_audit,
    build_parser,
    export_from_json,
    export_from_xlsx,
)


def test_build_annex_audit_marks_missing_cache() -> None:
    audit = build_annex_audit(None, cnae_normalized="4711302")
    assert audit.annex_status == "MISSING_CACHE"
    assert audit.annex_default is None


def test_export_from_xlsx_appends_annex_columns(db_session, tmp_path: Path) -> None:
    db_session.add(
        EconetCnaeCache(
            cnae="4711302",
            cnae_formatted="4711-3/02",
            description="Comercio varejista de mercadorias em geral",
            econet_id_cnae="econet-1",
            activity_types=["COMERCIO"],
            simples_status="ALLOWED",
            simples_allowed=True,
            simples_annex_default="I",
            simples_annex_conditional=None,
            factor_r_applicable=None,
            factor_r_threshold=None,
            mei_status="NOT_ALLOWED",
            mei_allowed=False,
            mei_occupation=None,
            presumed_profit_status="ALLOWED",
            presumed_profit_allowed=True,
            presumed_profit_irpj_rate=None,
            presumed_profit_csll_rate=None,
            actual_profit_status="UNKNOWN",
            actual_profit_mandatory=None,
            obligations_general={},
            obligations_simples={},
            obligations_simei={},
            unmapped_obligations=[],
            normalized_payload={},
            parse_status="PARSED",
            parser_version="econet-html-v2",
            content_hash="hash",
            retrieved_at=datetime.now(timezone.utc),
            expires_at=datetime.now(timezone.utc),
        )
    )
    db_session.flush()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalogo"
    sheet.append(("CNAE", "activity_type"))
    sheet.append(("4711-3/02", "COMERCIO"))
    input_path = tmp_path / "catalogo.xlsx"
    output_path = tmp_path / "catalogo_out.xlsx"
    workbook.save(input_path)

    summary = export_from_xlsx(db_session, input_path=input_path, output_path=output_path)
    assert summary["ok"] == 1

    result = load_workbook(output_path)
    catalog = result["Catalogo"]
    headers = [cell.value for cell in catalog[1]]
    assert "anexo_econet_padrao" in headers
    header_map = {header: index + 1 for index, header in enumerate(headers)}
    assert catalog.cell(row=2, column=header_map["anexo_econet_padrao"]).value == "I"
    assert catalog.cell(row=2, column=header_map["anexo_econet_status"]).value == "OK"
    assert "ResumoAnexos" in result.sheetnames


def test_export_from_json_creates_workbook(db_session, tmp_path: Path) -> None:
    db_session.add(
        EconetCnaeCache(
            cnae="5611201",
            cnae_formatted="5611-2/01",
            description="Restaurantes e similares",
            econet_id_cnae="econet-2",
            activity_types=["SERVICOS"],
            simples_status="ALLOWED",
            simples_allowed=True,
            simples_annex_default="I",
            simples_annex_conditional=None,
            factor_r_applicable=None,
            factor_r_threshold=None,
            mei_status="NOT_ALLOWED",
            mei_allowed=False,
            mei_occupation=None,
            presumed_profit_status="ALLOWED",
            presumed_profit_allowed=True,
            presumed_profit_irpj_rate=None,
            presumed_profit_csll_rate=None,
            actual_profit_status="UNKNOWN",
            actual_profit_mandatory=None,
            obligations_general={},
            obligations_simples={},
            obligations_simei={},
            unmapped_obligations=[],
            normalized_payload={},
            parse_status="PARSED",
            parser_version="econet-html-v2",
            content_hash="hash-2",
            retrieved_at=datetime.now(timezone.utc),
            expires_at=datetime.now(timezone.utc),
        )
    )
    db_session.flush()

    input_path = tmp_path / "catalogo.json"
    output_path = tmp_path / "catalogo.xlsx"
    input_path.write_text(
        json.dumps(
            {
                "catalog": [
                    {
                        "cnae": "5611-2/01",
                        "normalized": "5611201",
                        "division": "56",
                        "denomination": "Restaurantes e similares",
                        "activity_type": "SERVICOS",
                        "rule_id": "R_SERVICOS_FALLBACK",
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    summary = export_from_json(db_session, input_path=input_path, output_path=output_path)
    assert summary["ok"] == 1
    workbook = load_workbook(output_path)
    assert workbook["Catalogo"].cell(row=2, column=7).value == "I"


def test_parser_requires_single_input_source() -> None:
    args = build_parser().parse_args(["--input-xlsx", "a.xlsx"])
    assert str(args.input_xlsx).endswith("a.xlsx")


def test_script_help_runs() -> None:
    result = subprocess.run(
        [sys.executable, "backend/scripts/export_econet_simples_annex_audit.py", "--help"],
        capture_output=True,
        text=True,
        cwd=".",
    )
    assert result.returncode == 0
